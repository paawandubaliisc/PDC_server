

import time

# CT = time.time()
# SOC_SERVER = int(CT)
# FRACSEC_SERVER = (CT - SOC_SERVER)*10**6
# MILSEC_SERVER = int((time.time())*10**3)
# MILSEC_SERVER_calc = int(SOC_SERVER * 10**3 + FRACSEC_SERVER * 10**-3)


# arr = [1, 2, 3]
# print("arr[0] = ", arr[0])


# print("SOC_SERVER = ", SOC_SERVER)
# print("\nFRACSEC_SERVER = ", FRACSEC_SERVER)
# print("\nMILSEC_SERVER = ", MILSEC_SERVER)
# print("\nMILSEC_SERVER_calc = ", MILSEC_SERVER_calc)

# print("\nsend start time MILSEC_SERVER = ", MILSEC_SERVER + 600000)


from collections import deque

class queue:
    def __init__(self):
        self.queue = deque()

    def enqueue(self, data_recv):
        self.queue.appendleft(data_recv)
    
    def dequeue(self):
        return self.queue.pop()
    
    def size(self):
        return len(self.queue)
    
    def show_queue(self):
        print(self.queue)

set1 = ["sender2",1,2,3,4]
set2 = ["sender1",3,4,5,6]
#sender = "pdc"
import time
data_queue = queue()
t1 = time.perf_counter_ns()
data_queue.enqueue(set1)
data_queue.show_queue()

t2 = time.perf_counter_ns()
print("enqueue time is {}", format(t2-t1))
data_queue.enqueue(set2)
#print(data_queue.size())
data_queue.show_queue()
set3 = data_queue.dequeue()
print("set 3 is {}".format(set3))
# set3 = data_queue.dequeue()
# print(set3)


CT = time.time()
SOC_SERVER = int(CT)
FRACSEC_SERVER = (CT - SOC_SERVER)*10**6
MILSEC_SERVER = int((time.time())*10**3)
MILSEC_SERVER_calc = int(SOC_SERVER * 10**3 + FRACSEC_SERVER * 10**-3)
print("SOC_SERVER = ", SOC_SERVER)

print("\nFRACSEC_SERVER = ", FRACSEC_SERVER)
print("\nMILSEC_SERVER = ", MILSEC_SERVER)
print("\nMILSEC_SERVER_calc = ", MILSEC_SERVER_calc)

print("\nsend start time MILSEC_SERVER = ", MILSEC_SERVER + 200000)


'''

import openpyxl

# workbook = openpyxl.load_workbook("{}.xlsx".format(Test1))
workbook = openpyxl.load_workbook("Test1.xlsx")
sheet1 = workbook['{}'.format("testdb")]
row = sheet1.max_row
col = sheet1.max_column
print("total rows: {}".format(row))
print("total columns: {}".format(col))

for i in range(2,row):
        for j in range(1, col + 1):
            data = sheet1.cell(i,j).value
            print("row: {}, col: {}, value:{}".format(i,j,data))
'''

import socket

# socket = socket.gethostbyname(socket.gethostname())
# print(socket)
socket = "10.64.37.32"
dict = {'10.64.37.31' : 'Bus 08',
        '10.64.37.32' : 'Bus 26',
        '10.64.37.33' : 'Bus 17',
        '10.64.37.34' : 'Bus 30',
        '10.64.37.35' : 'pdc',
        '10.64.37.36' : 'Bus 38',
        '10.64.37.37' : 'Bus 37',
        '10.64.37.38' : 'Bus 65'}

print(dict[socket])