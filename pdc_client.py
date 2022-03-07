import socket
import time
import struct
import openpyxl
import logging



########################### Data Frame
SYNC_DATA = 0xAA00
STAT = 0xABCD
FRAME_SIZE = 48
ID_CODE = 1
VA = 0xDCBADCBADCBADCBA
VB = 0xDCBADCBADCBADCBA
VC = 0xDCBADCBADCBADCBA
IA = 0xDCBADCBADCBADCBA
IB = 0xDCBADCBADCBADCBA
IC = 0xDCBADCBADCBADCBA
FREQ = 0xABCD
DFREQ = 0xDCBA
ANALOG1 = 0xDCBADCBA
ANALOG2 = 0xDCBADCBA
ANALOG3 = 0xDCBADCBA
ANALOG4 = 0xDCBADCBA
DIGITAL = 0xABCD

local_ip = socket.gethostbyname(socket.gethostname())
# local_ip = "10.64.37.31"
dict = {'10.64.37.31' : 'Bus 08',
        '10.64.37.32' : 'Bus 26',
        '10.64.37.33' : 'Bus 17',
        '10.64.37.34' : 'Bus 30',
        '10.64.37.35' : 'pdc',
        '10.64.37.36' : 'Bus 38',
        '10.64.37.37' : 'Bus 37',
        '10.64.37.38' : 'Bus 65'}




def open_workbook(file_name, sheet_name):
    print(file_name, sheet_name)
    workbook = openpyxl.load_workbook("{}.xlsx".format(file_name))
    sheet1 = workbook['{}'.format(sheet_name)]
    row = sheet1.max_row
    col = sheet1.max_column
    print("Number of rows in file name {}, sheet {} is {} "
          .format(file_name, sheet_name, row))
    print("Number of columns in file name {}, sheet {} is {} "
          .format(file_name, sheet_name, col))
    return row, col, sheet1

def start_server(ip_addr,port_no):
    SERVER = (ip_addr)
    PORT = port_no
    SERVER_ADDR = (SERVER, PORT)
    BUFFER = 1024
    client = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    return SERVER_ADDR, client, BUFFER

def current_time():
    CT = time.time()
    SOC_CLIENT = int(CT)
    FRACSEC_CLIENT = int((CT - SOC_CLIENT)*10**6)
    return SOC_CLIENT, FRACSEC_CLIENT

def get_next_time():
    
    next_time = curr_time + 20
    #next_time = curr_time
    return next_time

def create_log(SERVER):
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.INFO)
    file_handler = logging.FileHandler('client.log')
    file_handler.setLevel(logging.INFO)
    formatter = logging.Formatter('%(asctime)s:%(levelname)s:%(message)s')
    logger.addHandler(file_handler)
    file_handler.setFormatter(formatter)
    logger.info('CLIENT STARTED at {}'.format(SERVER))
    return logger 

def start_sending(row, col, sheet1, client, logger, SERVER_ADDR, next_time):
    data = [0]*15
    for i in range(2,row):
        for j in range(1, col + 1):
            data[j] = sheet1.cell(i,j).value
        SOC_CLIENT, FRACSEC_CLIENT = current_time()
        # print("type of data 1 is {}", type(data[1]))
        # print("type of data 2 is {}", type(data[2]))
        # print("type of data 3 is {}", type(data[3]))
        # print("type of data 3 is {}", type(data[4]))
        # print("type of data 5 is {}", type(data[5]))
        # print("type of data 6 is {}", type(data[6]))
        # print("type of data 7 is {}", type(data[7]))
        # data[1] = float(data[1])
        msg = struct.pack('!3H2IH13d6Id',
                            SYNC_DATA,
                            FRAME_SIZE,
                            ID_CODE,
                            SOC_CLIENT,
                            FRACSEC_CLIENT,
                            STAT,
                            data[1],
                            data[3], data[4], data[5],
                            data[6], data[7], data[8],
                            data[9], data[10], data[11],
                            data[12], data[13], data[14],
                            FREQ, DFREQ,
                            ANALOG1, ANALOG2, ANALOG3, ANALOG4,
                            data[2])
        sent = 1
        while sent:
            if (next_time) - (int((time.time())*10**3)) < 2:
                client.sendto(msg, SERVER_ADDR)
                end3 = int((time.time())*10**3)
                logger.info("Message {} sent at {}".format(i-1, end3))
                next_time = next_time + 20
                sent = 0
            else:
                time.sleep(0.001)

def transmit_ON(file_name, sheet_name):
    print("Sheet name is : {}".format(sheet_name))
    SERVER_ADDR, client, BUFFER = start_server("10.64.37.35", 2345)
    row, col, sheet1 = open_workbook(file_name, sheet_name)
    logger = create_log(SERVER_ADDR)
    next_time = get_next_time()
    start_sending(row, col, sheet1, client, logger, SERVER_ADDR, next_time)
    print("end")


curr_time = 1646630865525
transmit_ON(file_name = "fault_on_next_line", sheet_name = dict[local_ip])


'''
print("\nSYNC FRAME is " + str(SYNC_DATA))
print("\nFRAME size is " + str(FRAME_SIZE))
print("\nPMU ID is " + str(ID_CODE))
print("\nSOC_CLIENT is " + str(SOC_CLIENT))
print("\nFRACSEC_CLIENT is " + str(FRACSEC_CLIENT))
print("\nSTAT is " + str(STAT))
print("\nVA is " + str(VA))
print("\nVB is " + str(VB))
print("\nVC is " + str(VC))
print("\nIA is " + str(IA))
print("\nIB is " + str(IB))
print("\nIC is " + str(IC))
print("\n\n")
'''
'''

second, microsec, millisec = current_time()
print("time is {}".format(time.time()))
print("\nseconds is {}".format(second))
print("\nmicrosec is {}".format(microsec))
print("\nmillisec is {}".format(millisec))
'''



'''
CT = time.time()
curr_time = int(CT*10**3)
next_time = curr_time + 20
print("\ncurr_time is {}".format(curr_time))
print("\nnext_time is {}".format(next_time))

#SOC_CLIENT = int(CT)
#FRACSEC_CLIENT = int((CT - SOC_CLIENT)*10**6)
#MILSEC_CLIENT = int((CT - SOC_CLIENT)*10**3)
#curr_time = SOC_CLIENT + MILSEC_CLIENT




'''
'''
i = 20
while i > 0:
    SOC_CLIENT = current_time()[0]
    FRACSEC_CLIENT = current_time()[1]
    msg = struct.pack('!HHHIIIHIHHHHIIIHH', 
                        SYNC_CONFIG, 
                        FRAME_SIZE, 
                        ID_CODE ,
                        SOC_CLIENT,
                        FRACSEC_CLIENT , 
                        TIME_BASE,
                        NUM_PMU, 
                        ST_NAME, 
                        FORMAT, 
                        PHNMR, 
                        ANNMR, 
                        DGNMR, 
                        CHNAM, 
                        PHUNIT, 
                        ANUNIT, 
                        DIGUNIT, 
                        FNOM)
    client.sendto(msg, SERVER_ADDR)
    print("\nSYNC is " + str(SYNC_CONFIG))
    print("\nFRAME size is " + str(FRAME_SIZE))
    print("\nPMU ID is " + str(ID_CODE))
    print("\nSOC_CLIENT is " + str(SOC_CLIENT))
    print("\nFRACSEC_CLIENT is " + str(FRACSEC_CLIENT))
    i = i - 1
'''
'''

###################### Config Frame
SYNC_CONFIG = 0xAA21
FRAME_SIZE = 48
ID_CODE = 1
TIME_BASE = 10**6
NUM_PMU = 0x1
ST_NAME = 0xABCD

FORMAT = 0x0007
PHNMR = 0x0006
vol_phas = 3
curr_phas = 3
ANNMR = 0x0002
DGNMR = 0x0000
CHNAM = 0xABCDABCD

PHUNIT = 0x10001000
ANUNIT = 0x02000032
DIGUNIT = 0x0
FNOM = 0x0001
'''

'''
#MILSEC_CLIENT = int((CT - SOC_CLIENT)*10**3)
    #print("Time is " + str(CT))
    #print("SOC Server in seconds is " + str(SOC_SERVER) + 
    #"\nFRACSEC in useconds is " + str(FRACSEC_SERVER))
'''